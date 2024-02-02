import openpyxl
import math
import sys

nameofinputfile = None
nameofoutputfile = None
mode = None
for param in sys.argv:
    if param[-4:] == ".txt" and mode == None:
        mode = "txttoxlsx"
        nameofinputfile = param
    elif param[-4:] == "xlsx" and mode == None:
        mode = "xlsxtotxt"
        nameofinputfile = param
    elif param[-4:] == ".txt":
        nameofoutputfile = param
    elif param[-4:] == "xlsx":
        nameofoutputfile = param

if nameofinputfile == None:
    print("The file is not supported!")
if nameofoutputfile == None:
    if mode == "txttoxlsx":
        nameofoutputfile = "output.xlsx"
    elif mode == "xlsxtotxt":
        nameofoutputfile = "output.txt"


if mode == "txttoxlsx":
    inputFile = open(nameofinputfile, "r")
    outputFile = openpyxl.Workbook()

    lines = inputFile.readlines()
    x = 1
    ws = outputFile.active

    redFill = openpyxl.styles.PatternFill(start_color='ff6b6b',
                    end_color='ff6b6b',
                    fill_type='solid')
    y = 1
    for line in lines:
        print(f"{math.floor(y / (len(lines) / 100))}% ..... {y} | {len(lines)}                                                                                    ", end="\r")
        y = y + 1
        line = line.replace(",", " ")
        splitLine = line.split()
        if len(splitLine) <= 1:
            continue 
        elif splitLine[0][:2] == "//":
            continue
        elif splitLine[0] == "<CurrencyName>":
            ws = outputFile.create_sheet(f"{splitLine[0]} {splitLine[1]}")
        elif splitLine[0] == "<Currency>":
            ws.cell(row=x, column=1, value=splitLine[1])
            ws.cell(row=x, column=2, value=splitLine[2])
            x = x + 1
        elif splitLine[0] == "<Trader>":
            nameofsheet = ""
            for pn in splitLine[1:]:
                if pn == "//":
                    break
                nameofsheet = nameofsheet + pn + " "
            ws = outputFile.create_sheet(nameofsheet)
            x = 1
            ws.cell(row=1, column=1, value="Имя")
            ws.cell(row=1, column=2, value="Тип")
            ws.cell(row=1, column=3, value="Купля")
            ws.cell(row=1, column=4, value="Продажа")

        elif splitLine[0] == "<Category>":
            ws.cell(row=x, column=1, value="Категория").fill = redFill
            nameofcategory = ""
            for o in splitLine[1:]:
                nameofcategory = nameofcategory + o + " "
            ws.cell(row=x, column=2, value=nameofcategory).fill = redFill
            ws.cell(row=x, column=3, value="1").fill = redFill
            ws.cell(row=x, column=4, value="60").fill = redFill
            
            categoryx = x+1
            x = x + 1
        else:
            ws.cell(row=x, column=1, value=splitLine[0])
            ws.cell(row=x, column=2, value=splitLine[1])
            ws.cell(row=x, column=3, value=splitLine[2])
            ws.cell(row=x, column=4, value=splitLine[3]).fill = redFill
            x = x + 1

    outputFile.save(nameofoutputfile)

elif mode == "xlsxtotxt":
    inputFile = openpyxl.load_workbook(nameofinputfile)
    outputFile = open(nameofoutputfile, "w+")

    sheets = inputFile.sheetnames
    sheets.remove("Sheet")

    for x in sheets:
        sheet = inputFile[x]
        outputFile.write(f"\n<Trader> {x}\n")
        print(x)
        y = 1
        mod = 1
        soldpercent = 60
        while True:
            if sheet.cell(row=y, column=1).value == "Категория":
                outputFile.write(f"\t<Category> {sheet.cell(row=y, column=2).value}\n")
                mod = float(sheet.cell(row=y, column=3).value)
                soldpercent = float(sheet.cell(row=y, column=4).value)
            elif sheet.cell(row=y, column=1).value == None:
                break
            else:
                outputFile.write(f"\t\t{sheet.cell(row=y, column=1).value},\t{sheet.cell(row=y, column=2).value},\t{math.ceil(int(sheet.cell(row=y, column=3).value) * mod)},\t{math.ceil((int(sheet.cell(row=y, column=3).value) / 100) * soldpercent)}\n")
            y = y + 1