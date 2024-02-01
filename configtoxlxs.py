import openpyxl

inputFile = open("input.txt", "r")
outputFile = openpyxl.Workbook()

lines = inputFile.readlines()
x = 1
ws = outputFile.active

redFill = openpyxl.styles.PatternFill(start_color='ff6b6b',
                   end_color='ff6b6b',
                   fill_type='solid')

for line in lines:
    line = line.replace(",", " ")
    splitLine = line.split()
    if len(splitLine) <= 1:
        continue 
    elif splitLine[0] == "<Trader>":
        nameofsheet = ""
        for pn in splitLine[1:]:
            nameofsheet = nameofsheet + pn + " "
        ws = outputFile.create_sheet(nameofsheet)
        x = 1
        ws.cell(row=1, column=1, value="Имя")
        ws.cell(row=1, column=2, value="Тип")
        ws.cell(row=1, column=3, value="Купля")
        ws.cell(row=1, column=4, value="Продажа")

    elif splitLine[0] == "<Category>":
        ws.cell(row=x, column=1, value="Категория").fill = redFill
        nameofcategory = ""
        for o in splitLine[1:]:
            nameofcategory = nameofcategory + o + " "
        ws.cell(row=x, column=2, value=nameofcategory).fill = redFill
        ws.cell(row=x, column=3, value="1").fill = redFill
        ws.cell(row=x, column=4, value="60").fill = redFill
        
        categoryx = x+1
        x = x + 1
    else:
        ws.cell(row=x, column=1, value=splitLine[0])
        ws.cell(row=x, column=2, value=splitLine[1])
        ws.cell(row=x, column=3, value=splitLine[2])
        ws.cell(row=x, column=4, value=splitLine[3]).fill = redFill
        x = x + 1

outputFile.save("Output.xlsx")