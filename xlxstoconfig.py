import openpyxl
import math

inputFile = openpyxl.load_workbook('input.xlsx')
outputFile = open("output.txt", "w+")

sheets = inputFile.sheetnames
sheets.remove("Sheet")

for x in sheets:
    sheet = inputFile[x]
    outputFile.write(f"\n<Trader> {x}\n")
    print(x)
    y = 1
    mod = 1
    soldpercent = 60
    while True:
        if sheet.cell(row=y, column=1).value == "Категория":
            outputFile.write(f"\t<Category> {sheet.cell(row=y, column=2).value}\n")
            mod = float(sheet.cell(row=y, column=3).value)
            soldpercent = float(sheet.cell(row=y, column=4).value)
        elif sheet.cell(row=y, column=1).value == None:
            break
        else:
            outputFile.write(f"\t\t{sheet.cell(row=y, column=1).value},\t{sheet.cell(row=y, column=2).value},\t{math.ceil(int(sheet.cell(row=y, column=3).value) * mod)},\t{math.ceil((int(sheet.cell(row=y, column=3).value) / 100) * soldpercent)}\n")
        y = y + 1